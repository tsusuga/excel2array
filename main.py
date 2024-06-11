from lxml import etree
from openpyxl import load_workbook
import os
import re

# folder_path ='folder/'

# files = os.listdir(folder_path)

# for file in files:
#   full_path = os.path.join(folder_path, file)

#   if os.path.isfile(full_path):
#     print(file)
#     with open(full_path, 'a', encoding='utf-8' ) as f:
#       f.write('hello')


# def excel_to_xml(file_path):
#   wb = load_workbook(file_path) # Load the Excel file
#   xml_data = []

#   print(wb.sheetnames)
#   for sheet_name in wb.sheetnames: #シートごとに処理？
#     ws = wb[sheet_name]
#     root = etree.Element("sheet_name")

#     for row in ws.iter_rows(values_only=True):
#       xml_row = etree.SubElement(root, 'row')
#       for value in row:
#         etree.SubElement(xml_row, 'cell').text = str(value)

#     xml_data.append(etree.tostring(root, pretty_print=True).decode('utf-8'))

#   return xml_data


excel_file = "data.xlsx"

# xml_data_list = excel_to_xml(excel_file)

# for i, xml_data in enumerate(xml_data_list, start=1):
#   with open(f"dest/data{i}.txt", "w", newline="") as f:
#     f.write(xml_data)

wb = load_workbook(excel_file)
ws = wb['Sheet3']

arrayData = []

for row in ws.iter_rows(values_only=True):
  arrayData.append(row)

with open('dest/output.txt', 'w', encoding='utf-8' ) as f:
  for row in arrayData:
    f.write(str(row) + '\n')

# with open('dest/output.txt', 'r', encoding='utf-8' ) as f:
#   data = f.read()

# oldString1 = r'\(\''
# oldString2 = r'\)\''
# newString1 = '[\''
# newString2 = '\']'

# data = re.sub(oldString1, newString1, data)
# data = re.sub(oldString2, newString2, data)

# with open('dest/output.txt', 'w', encoding='utf-8' ) as f:
#   f.write(data)
